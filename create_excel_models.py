#!/usr/bin/env python3
"""
Script para criar arquivos modelo Excel para importação
"""
import sys
sys.path.append('/usr/lib/python3/dist-packages')

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from datetime import date
except ImportError:
    print("Instalando openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "--user", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from datetime import date


def create_proprietarios_model():
    """Cria modelo Excel para proprietários - Formato Avançado"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Proprietarios"
    
    # Cabeçalhos conforme especificado no prompt
    headers = ['Nome', 'Sobrenome', 'Documento', 'Tipo Documento', 'Endereço', 'Telefone', 'Email']
    ws.append(headers)
    
    # Estilizar cabeçalhos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Exemplos conforme especificado
    ws.append(['João', 'Silva', '170.858.698-95', 'CPF', 'Rua das Flores, 123', '(11) 98765-4321', 'joao@email.com'])
    ws.append(['Maria', 'Santos', '987.654.321-00', 'CPF', 'Av. Paulista, 456', '(11) 91234-5678', 'maria@email.com'])
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 25
    
    wb.save('Proprietarios.xlsx')
    print("✓ Proprietarios.xlsx criado (formato avançado)")


def create_imoveis_model():
    """Cria modelo Excel para imóveis - Formato Avançado"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Imoveis"
    
    # Cabeçalhos conforme especificado no prompt
    headers = ['Nome', 'Endereço', 'Tipo', 'Área Total', 'Área Construída', 'Valor Catastral', 'Valor Mercado', 'IPTU Anual', 'Condomínio']
    ws.append(headers)
    
    # Estilizar cabeçalhos
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Exemplos conforme especificado
    ws.append([
        'Casa Centro', 'Rua das Flores, 123', 'Residencial', '150,50', '140,00', '200.000,00', '350.000,00', '1.200,00', ''
    ])
    ws.append([
        'Sala Comercial', 'Av. Paulista, 456', 'Comercial', '85,00', '85,00', '', '500.000,00', '3.500,00', '800,00'
    ])
    
    # Ajustar largura das colunas
    for i, width in enumerate([20, 30, 15, 15, 18, 18, 18, 15, 15], start=1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    wb.save('Imoveis.xlsx')
    print("✓ Imoveis.xlsx criado (formato avançado)")


def create_participacoes_model():
    """Cria modelo Excel para participações - Formato Especial"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Participacoes"
    
    # Cabeçalhos especiais: Nome, Endereço, VALOR, [proprietários dinâmicos]
    headers = ['Nome', 'Endereço', 'VALOR', 'João Silva', 'Maria Santos']
    ws.append(headers)
    
    # Estilizar cabeçalhos
    header_fill = PatternFill(start_color="E74856", end_color="E74856", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Exemplos - VALOR sempre 100,000000 %
    ws.append(['Casa Centro', 'Rua das Flores, 123', '100,000000 %', '60,000000 %', '40,000000 %'])
    ws.append(['Sala Comercial', 'Av. Paulista, 456', '100,000000 %', '100,000000 %', ''])
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    # Adicionar instruções
    ws['A4'] = 'IMPORTANTE:'
    ws['A5'] = '1. A coluna VALOR deve sempre ser "100,000000 %"'
    ws['A6'] = '2. A soma dos percentuais deve ser exatamente 100%'
    ws['A7'] = '3. Adicione uma coluna para cada proprietário'
    ws['A8'] = '4. Use o formato "XX,XXXXXX %" para os percentuais'
    
    for row in range(4, 9):
        ws[f'A{row}'].font = Font(bold=True, color="FF0000")
    
    wb.save('Participacoes.xlsx')
    print("✓ Participacoes.xlsx criado (formato especial)")


def create_alugueis_model():
    """Cria modelo Excel para aluguéis - MÚLTIPLAS PLANILHAS"""
    wb = Workbook()
    
    # Remover planilha padrão
    wb.remove(wb.active)
    
    # Criar uma planilha por mês (exemplo com 3 meses)
    meses_exemplo = [
        ('Setembro_2024', '24/09/2024'),
        ('Outubro_2024', '24/10/2024'),
        ('Novembro_2024', '24/11/2024')
    ]
    
    for mes_nome, data_ref in meses_exemplo:
        ws = wb.create_sheet(mes_nome)
        
        # Data de referência na célula A1
        ws['A1'] = data_ref
        ws['A1'].font = Font(bold=True, size=12)
        
        # Cabeçalhos: Imóvel, Valor Total, [Proprietários], Taxa Admin
        headers = ['Imóvel/Endereço', 'Valor Total', 'João Silva', 'Maria Santos', 'Taxa Administração']
        for i, header in enumerate(headers, 1):
            ws.cell(row=2, column=i, value=header)
        
        # Estilizar cabeçalhos
        header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        header_font = Font(bold=True, color="000000")
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=2, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Exemplos de dados
        exemplos = [
            ['Casa Centro', '2.500,00', '1.500,00', '1.000,00', '50,00'],
            ['Sala Comercial', '3.000,00', '3.000,00', '', '75,00'],
            ['Apartamento Sul', '- 1.200,00', '- 1.200,00', '', '30,00']  # Valor negativo
        ]
        
        for row_idx, exemplo in enumerate(exemplos, 3):
            for col_idx, valor in enumerate(exemplo, 1):
                ws.cell(row=row_idx, column=col_idx, value=valor)
        
        # Ajustar largura das colunas
        for i, width in enumerate([25, 15, 15, 15, 18], start=1):
            ws.column_dimensions[chr(64 + i)].width = width
        
        # Adicionar instruções na planilha
        ws['A8'] = 'INSTRUÇÕES:'
        ws['A9'] = '1. A célula A1 contém a data de referência (DD/MM/YYYY)'
        ws['A10'] = '2. Valores negativos usam hífen: "- 1.200,00"'
        ws['A11'] = '3. Soma dos valores individuais + taxa = valor total'
        ws['A12'] = '4. Uma planilha por mês'
        
        for row in range(8, 13):
            ws[f'A{row}'].font = Font(bold=True, color="FF0000")
    
    wb.save('Alugueis.xlsx')
    print("✓ Alugueis.xlsx criado (múltiplas planilhas)")


if __name__ == "__main__":
    print("\n🔨 Criando arquivos modelo Excel...\n")
    
    create_proprietarios_model()
    create_imoveis_model()
    create_participacoes_model()
    create_alugueis_model()
    
    print("\n✅ Todos os modelos foram criados com sucesso!")
    print("\nArquivos criados:")
    print("  - Proprietarios.xlsx")
    print("  - Imoveis.xlsx")
    print("  - Alugueis.xlsx")
    print("  - Participacoes.xlsx")
    print("\nUse estes arquivos como modelo para importar dados no sistema.\n")
